#!/usr/bin/env python3
"""
PDFcsv - Universal PDF to CSV Extractor
========================================

A powerful CLI tool for extracting tabular data from PDF documents with
intelligent column detection, bank statement support, and interactive
structure selection.

Features:
    - Gap-based column detection using character X-positions
    - Bank statement support with debit/credit column filling
    - Interactive CLI with arrow-key navigation
    - Structure grouping by column positions
    - Multi-language keyword detection

Usage:
    python pdfcsv.py input.pdf              # Interactive mode
    python pdfcsv.py input.pdf --analyze    # Analyze column structure
    python pdfcsv.py input.pdf --columns 6  # Extract specific column count
    python pdfcsv.py input.pdf -o out.csv   # Custom output path
    python pdfcsv.py input.pdf --gap 10     # Adjust gap threshold

Requirements:
    pip install pdfplumber

Author: @stexz01
License: MIT
Repository: https://github.com/stexz01/pdfcsv
"""

__version__ = "1.5.0"
__author__ = "@stexz01"

# Global silent mode flag
SILENT_MODE = False

# ═══════════════════════════════════════════════════════════════════════════════
# IMPORTS
# ═══════════════════════════════════════════════════════════════════════════════

import csv
import json
import os
import re
import sys
from collections import Counter

import pdfplumber

# Optional: Banking keywords for statement detection (graceful fallback if missing)
try:
    from banking_keywords import is_bank_statement, DEBIT_KEYWORDS, CREDIT_KEYWORDS
    BANKING_DETECTION_AVAILABLE = True
except ImportError:
    BANKING_DETECTION_AVAILABLE = False

# Terminal raw input (Unix only, graceful fallback for Windows)
try:
    import tty
    import termios
    _TERMINAL_RAW_INPUT = True
except ImportError:
    _TERMINAL_RAW_INPUT = False


# ═══════════════════════════════════════════════════════════════════════════════
# COLORS & STYLING
# ═══════════════════════════════════════════════════════════════════════════════

class Colors:
    # Basic colors
    RED = '\033[91m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    MAGENTA = '\033[95m'
    CYAN = '\033[96m'
    WHITE = '\033[97m'
    GRAY = '\033[90m'

    # Styles
    BOLD = '\033[1m'
    DIM = '\033[2m'
    UNDERLINE = '\033[4m'

    # Reset
    RESET = '\033[0m'

    @classmethod
    def disable(cls):
        """Disable colors for non-TTY output"""
        for attr in dir(cls):
            if attr.isupper():
                setattr(cls, attr, '')


# Disable colors if not a terminal
if not sys.stdout.isatty():
    Colors.disable()


def banner():
    """Print the PDFcsv banner"""
    c = Colors
    print(f"""
{c.CYAN}{c.BOLD}╔═══════════════════════════════════════════════════════════════╗
║                                                               ║
║   {c.WHITE}██████╗ ██████╗ ███████╗{c.YELLOW}     ██████╗███████╗██╗   ██╗{c.CYAN}      ║
║   {c.WHITE}██╔══██╗██╔══██╗██╔════╝{c.YELLOW}    ██╔════╝██╔════╝██║   ██║{c.CYAN}      ║
║   {c.WHITE}██████╔╝██║  ██║█████╗  {c.YELLOW}    ██║     ███████╗██║   ██║{c.CYAN}      ║
║   {c.WHITE}██╔═══╝ ██║  ██║██╔══╝  {c.YELLOW}    ██║     ╚════██║╚██╗ ██╔╝{c.CYAN}      ║
║   {c.WHITE}██║     ██████╔╝██║     {c.YELLOW}    ╚██████╗███████║ ╚████╔╝ {c.CYAN}      ║
║   {c.WHITE}╚═╝     ╚═════╝ ╚═╝     {c.YELLOW}     ╚═════╝╚══════╝  ╚═══╝  {c.CYAN}      ║
║                                                               ║
║   {c.DIM}Universal PDF to CSV Extractor {c.WHITE}v{__version__}{c.CYAN}                       ║
║   {c.DIM}Made with {c.RED}♥{c.DIM} by {c.WHITE}{__author__}{c.CYAN}                                    ║
║                                                               ║
╚═══════════════════════════════════════════════════════════════╝{c.RESET}
""")


def print_step(num, text):
    """Print a step indicator"""
    if SILENT_MODE:
        return
    c = Colors
    print(f"{c.CYAN}{c.BOLD}[{num}]{c.RESET} {text}")


def print_success(text):
    """Print success message"""
    if SILENT_MODE:
        return
    c = Colors
    print(f"{c.GREEN}{c.BOLD}✓{c.RESET} {text}")


def print_error(text):
    """Print error message (always shown, even in silent mode)"""
    c = Colors
    print(f"{c.RED}{c.BOLD}✗{c.RESET} {text}")


def print_info(text):
    """Print info message"""
    if SILENT_MODE:
        return
    c = Colors
    print(f"{c.BLUE}ℹ{c.RESET} {text}")


def print_warning(text):
    """Print warning message"""
    if SILENT_MODE:
        return
    c = Colors
    print(f"{c.YELLOW}⚠{c.RESET} {text}")


# ═══════════════════════════════════════════════════════════════════════════════
# CORE EXTRACTION LOGIC
# ═══════════════════════════════════════════════════════════════════════════════

def open_pdf(pdf_path: str):
    """
    Open a PDF file with automatic handling of encrypted documents.

    Attempts to open normally first, then tries empty password for PDFs
    with owner-only restrictions. If password-protected, prompts user
    for password with retry support.

    Args:
        pdf_path: Path to the PDF file

    Returns:
        pdfplumber.PDF object

    Raises:
        SystemExit: If user cancels password entry
    """
    c = Colors

    try:
        return pdfplumber.open(pdf_path)
    except Exception as e1:
        # Try with empty password for encrypted PDFs
        try:
            return pdfplumber.open(pdf_path, password="")
        except Exception:
            # Check if it's an encryption error (check str, repr, and type)
            error_info = f"{str(e1).lower()} {repr(e1).lower()} {str(type(e1))}"
            if "password" in error_info or "encrypt" in error_info:
                print(f"\n{c.YELLOW}{c.BOLD}PDF is password-protected{c.RESET}")
                print(f"{c.DIM}File: {pdf_path}{c.RESET}")
                print(f"{c.DIM}Enter password or 'q' to quit{c.RESET}\n")

                # Password retry loop
                attempts = 0
                while True:
                    try:
                        attempts += 1
                        password = input(f"{c.CYAN}Password:{c.RESET} ")

                        # Check for quit
                        if password.lower() == 'q':
                            print(f"{c.YELLOW}Cancelled{c.RESET}\n")
                            sys.exit(0)

                        # Try opening with password
                        try:
                            pdf = pdfplumber.open(pdf_path, password=password)
                            print(f"{c.GREEN}{c.BOLD}✓{c.RESET} Password accepted\n")
                            return pdf
                        except Exception:
                            print(f"{c.RED}✗ Wrong password{c.RESET} {c.DIM}(attempt {attempts}){c.RESET}")
                            continue

                    except (KeyboardInterrupt, EOFError):
                        print(f"\n{c.YELLOW}Cancelled{c.RESET}\n")
                        sys.exit(0)
            raise e1


def _extract_tables(pdf) -> list:
    """
    Extract data from an open PDF using pdfplumber's built-in table detection.

    Works for PDFs with structured tables (gridlines/borders).
    Handles multi-line cells by joining text within each cell.

    Args:
        pdf: An open pdfplumber.PDF object

    Returns:
        List of rows (each row is a list of strings), or empty list if no tables found.
    """
    all_rows = []
    for page in pdf.pages:
        tables = page.extract_tables()
        for table in tables:
            for row in table:
                cleaned = []
                for cell in row:
                    if cell:
                        text = re.sub(r'\s+', ' ', cell.replace('\n', ' ').strip())
                        cleaned.append(text)
                    else:
                        cleaned.append('')
                all_rows.append(cleaned)
    return all_rows


def clean_table_rows(rows: list, min_filled: int = 2) -> list:
    """
    Clean extracted table rows:
    - Remove fully empty rows
    - Remove rows with too few meaningful values (page-break header remnants)

    Args:
        rows: Raw table rows
        min_filled: Minimum non-empty, non-dash cells to keep a row

    Returns:
        Cleaned list of rows
    """
    if not rows:
        return []

    cleaned = []
    for row in rows:
        meaningful = sum(
            1 for cell in row
            if cell and cell.strip() and cell.strip() != '-'
        )
        if meaningful >= min_filled:
            cleaned.append(row)
    return cleaned


def _table_has_concatenated_cells(rows: list) -> bool:
    """
    Detect if table rows contain concatenated multi-record data.

    Some PDFs (e.g. HDFC statements) pack multiple transactions into a single
    table cell. This produces rows where date/amount fields contain multiple
    space-separated values, which is not useful for CSV extraction.

    Returns True if cells appear to contain concatenated records.
    """
    if len(rows) < 2:
        return False

    # Check first few data rows (skip possible header)
    data_rows = rows[1:min(6, len(rows))]
    if not data_rows:
        data_rows = rows[:min(5, len(rows))]

    for row in data_rows:
        for cell in row:
            if not cell or len(cell) < 20:
                continue
            # Multiple date patterns in a single cell = concatenated records
            if len(re.findall(r'\d{2}/\d{2}/\d{2,4}', cell)) > 1:
                return True
            if len(re.findall(r'\d{1,2}\s+\w{3}\s+\d{2,4}', cell)) > 1:
                return True

    return False


def _warn_broken_headers(raw_rows: list, cleaned_rows: list):
    """
    Detect and warn if column headers are missing from table extraction.

    Some PDFs (e.g. SBI statements) render header text as vector graphics
    (tiny filled rectangles) instead of actual text characters. This makes
    them invisible to any text-based PDF extraction library.

    Checks raw table rows for rows that were cleaned away (mostly empty cells)
    which likely represent broken/unextractable header rows.
    """
    if not raw_rows or not cleaned_rows:
        return

    # Find the dominant column count in cleaned data
    col_counts = Counter(len(r) for r in cleaned_rows)
    target_cols = col_counts.most_common(1)[0][0]

    # Check raw rows that were removed by cleaning — they might be broken headers
    cleaned_set = {tuple(r) for r in cleaned_rows}
    for row in raw_rows:
        if tuple(row) in cleaned_set or len(row) != target_cols:
            continue

        # Row was removed and has the right column count — check if it's a broken header
        empty_count = sum(1 for cell in row if not cell or not cell.strip())
        if empty_count > len(row) // 2:
            # More than half the cells are empty — broken header
            filled = [cell.strip() for cell in row if cell and cell.strip()]
            filled_str = ', '.join(filled) if filled else 'none'
            print_warning(
                f"Column headers not extractable (rendered as vector graphics in PDF)"
            )
            print_info(
                f"{Colors.DIM}Only partial header recovered: [{filled_str}]{Colors.RESET}"
            )
            return


def table_rows_to_lines(rows: list) -> list:
    """
    Convert table-extracted rows to the lines format used by gap-based extraction.

    Uses dummy x_positions since table extraction doesn't provide pixel positions.
    """
    return [
        {
            'col_count': len(row),
            'columns': row,
            'x_positions': list(range(len(row)))
        }
        for row in rows
    ]


def _extract_gaps(pdf, gap_threshold: int = 5) -> list:
    """
    Internal: gap-based column extraction from an open pdfplumber PDF.

    Groups characters by Y position, detects column breaks where gaps
    between characters exceed threshold.
    """
    all_lines = []

    for page in pdf.pages:
        chars = page.chars
        if not chars:
            continue

        # Group characters by Y position (lines)
        lines_by_y = {}
        for c in chars:
            y = round(c['top'])
            if y not in lines_by_y:
                lines_by_y[y] = []
            lines_by_y[y].append(c)

        # Process each line
        for y in sorted(lines_by_y.keys()):
            row_chars = sorted(lines_by_y[y], key=lambda c: c['x0'])

            # Build columns by detecting gaps
            columns = []
            x_positions = []
            current_col = []
            col_start_x = None

            for i, char in enumerate(row_chars):
                if i > 0:
                    gap = char['x0'] - row_chars[i-1]['x1']
                    if gap > gap_threshold:
                        col_text = ''.join(c['text'] for c in current_col)
                        if col_text.strip():
                            columns.append(col_text.strip())
                            x_positions.append(col_start_x)
                        current_col = []
                        col_start_x = None

                if col_start_x is None:
                    col_start_x = char['x0']
                current_col.append(char)

            # Don't forget last column
            if current_col:
                col_text = ''.join(c['text'] for c in current_col)
                if col_text.strip():
                    columns.append(col_text.strip())
                    x_positions.append(col_start_x)

            if columns:
                all_lines.append({
                    'col_count': len(columns),
                    'columns': columns,
                    'x_positions': x_positions
                })

    return all_lines


def extract_lines_with_gaps(pdf_path: str, gap_threshold: int = 5) -> list:
    """
    Extract text lines from PDF using gap-based column detection.

    This is the core extraction algorithm. It works by:
    1. Extracting all characters with their X,Y positions
    2. Grouping characters by Y position (same line)
    3. Detecting column breaks where gaps between characters exceed threshold
    4. Recording X positions for structure-based grouping

    Args:
        pdf_path: Path to the PDF file
        gap_threshold: Minimum pixel gap to consider a column break (default: 5)

    Returns:
        List of dicts, each containing:
            - col_count: Number of columns detected
            - columns: List of text values for each column
            - x_positions: List of X coordinates for each column start
    """
    with open_pdf(pdf_path) as pdf:
        return _extract_gaps(pdf, gap_threshold)


def find_header_row(lines: list) -> dict:
    """
    Find the header row in bank statement data.

    Searches for rows containing both debit and credit keywords,
    preferring rows with more columns (typical of headers).

    Args:
        lines: List of extracted line dicts

    Returns:
        Header line dict or None if not found
    """
    if not BANKING_DETECTION_AVAILABLE:
        return None

    best_match = None
    best_score = 0

    for line in lines:
        text = ' '.join(line['columns']).lower()
        score = 0

        # Check for debit keywords
        for kw in DEBIT_KEYWORDS:
            if kw in text:
                score += 1
                break

        # Check for credit keywords
        for kw in CREDIT_KEYWORDS:
            if kw in text:
                score += 1
                break

        # Prefer rows with more columns (likely headers)
        if score >= 2 and line['col_count'] > best_score:
            best_match = line
            best_score = line['col_count']

    return best_match


def keyword_matches(col: str, keywords: set) -> bool:
    """
    Check if column header matches banking keywords with word boundary awareness.

    Uses smart matching to prevent false positives:
    - Short keywords (<=3 chars): require word boundary match
    - Longer keywords: substring match is sufficient

    Example: 'cr' won't match 'description' but will match 'CR' or 'Cr.'

    Args:
        col: Column header text
        keywords: Set of keywords to match against

    Returns:
        True if any keyword matches
    """
    col_lower = col.lower().strip()

    for kw in keywords:
        # For very short keywords (2-3 chars), require exact match or word boundary
        if len(kw) <= 3:
            # Check if column equals keyword exactly, or keyword is at word boundary
            if col_lower == kw or re.search(rf'\b{re.escape(kw)}\b', col_lower):
                return True
        else:
            # For longer keywords, substring match is fine
            if kw in col_lower:
                return True
    return False


def find_debit_credit_column_indices(header: dict) -> tuple:
    """
    Identify debit and credit column positions in a bank statement header.

    Args:
        header: Header line dict with 'columns' key

    Returns:
        Tuple of (debit_index, credit_index), or (None, None) if not found
    """
    if not header or not BANKING_DETECTION_AVAILABLE:
        return None, None

    debit_idx = None
    credit_idx = None

    for i, col in enumerate(header['columns']):
        if debit_idx is None and keyword_matches(col, DEBIT_KEYWORDS):
            debit_idx = i
        if credit_idx is None and keyword_matches(col, CREDIT_KEYWORDS):
            credit_idx = i

    return debit_idx, credit_idx


def fill_empty_columns(lines: list, header: dict, target_col_count: int) -> list:
    """
    Fill missing debit/credit columns with '-' using gap analysis.

    Bank statements often have empty debit OR credit columns per row.
    This function determines which column is missing by comparing the
    gap between the amount and balance columns against header patterns.

    Handles:
        - 1 missing column (debit OR credit empty)
        - 2 missing columns (both debit AND credit need filling)

    Args:
        lines: List of extracted line dicts
        header: Header row dict with column positions
        target_col_count: Expected column count for data rows

    Returns:
        List of column lists with '-' inserted where needed
    """
    if not header or 'x_positions' not in header:
        return [line['columns'] for line in lines if line['col_count'] == target_col_count]

    header_x = header['x_positions']
    header_cols = header['columns']
    debit_idx, credit_idx = find_debit_credit_column_indices(header)

    if debit_idx is None or credit_idx is None:
        return [line['columns'] for line in lines if line['col_count'] == target_col_count]

    # Balance column is the last one
    balance_idx = len(header_cols) - 1

    # Calculate expected gaps from header positions
    header_debit_to_balance = header_x[balance_idx] - header_x[debit_idx]
    header_credit_to_balance = header_x[balance_idx] - header_x[credit_idx]

    filled_rows = []

    for line in lines:
        # If line has same column count as header, use as-is
        if line['col_count'] == len(header_cols):
            filled_rows.append(line['columns'])
            continue

        # If line has one less column than header (missing debit OR credit)
        if line['col_count'] == len(header_cols) - 1:
            columns = list(line['columns'])
            x_pos = line.get('x_positions', [])

            if not x_pos or len(x_pos) < 2:
                filled_rows.append(line['columns'])
                continue

            amount_col_idx = debit_idx
            balance_col_idx = len(columns) - 1

            if amount_col_idx < len(x_pos) and balance_col_idx < len(x_pos):
                amount_x = x_pos[amount_col_idx]
                balance_x = x_pos[balance_col_idx]
                data_gap = balance_x - amount_x

                diff_to_debit_pattern = abs(data_gap - header_debit_to_balance)
                diff_to_credit_pattern = abs(data_gap - header_credit_to_balance)

                if diff_to_debit_pattern < diff_to_credit_pattern:
                    columns.insert(credit_idx, '-')
                else:
                    columns.insert(debit_idx, '-')

                filled_rows.append(columns)
            else:
                filled_rows.append(line['columns'])

        # If line has TWO less columns than header (missing Chq No AND Debit/Credit)
        # Common pattern: [Date, Particulars, Amount, Balance, Init] → need [Date, Chq, Part, Debit, Credit, Bal, Init]
        elif line['col_count'] == len(header_cols) - 2:
            columns = list(line['columns'])
            x_pos = line.get('x_positions', [])

            if not x_pos or len(x_pos) < 2:
                filled_rows.append(line['columns'])
                continue

            # Step 1: Insert '-' for missing Chq No (usually at index 1)
            # This assumes the pattern: Date is col 0, then Chq No is missing, Particulars is col 1 in data
            chq_idx = 1  # Chq No is typically column 1 in bank statements
            if chq_idx < debit_idx:  # Only insert if Chq comes before Debit/Credit
                columns.insert(chq_idx, '-')

            # Now columns has header-1 count, apply normal debit/credit gap-fill
            # Recalculate indices after insertion
            amount_col_idx = debit_idx  # In the now 6-col row
            balance_col_idx = len(columns) - 2  # Second to last (before Init)

            if amount_col_idx < len(x_pos) and balance_col_idx < len(x_pos):
                # Use original x_pos (before Chq insertion) to determine gap
                orig_amount_idx = amount_col_idx - 1  # Adjust for inserted Chq
                orig_balance_idx = balance_col_idx - 1

                if orig_amount_idx >= 0 and orig_amount_idx < len(x_pos) and orig_balance_idx < len(x_pos):
                    amount_x = x_pos[orig_amount_idx]
                    balance_x = x_pos[orig_balance_idx]
                    data_gap = balance_x - amount_x

                    diff_to_debit = abs(data_gap - header_debit_to_balance)
                    diff_to_credit = abs(data_gap - header_credit_to_balance)

                    # Insert '-' for missing Debit or Credit
                    if diff_to_debit < diff_to_credit:
                        # Amount is in debit position, credit is missing
                        columns.insert(credit_idx, '-')
                    else:
                        # Amount is in credit position, debit is missing
                        columns.insert(debit_idx, '-')

                    filled_rows.append(columns)
                else:
                    # Fallback: just insert both placeholders
                    if credit_idx > debit_idx:
                        columns.insert(credit_idx, '-')
                    else:
                        columns.insert(debit_idx, '-')
                    filled_rows.append(columns)
            else:
                # Can't determine gap, insert credit placeholder as default
                columns.insert(credit_idx, '-')
                filled_rows.append(columns)

        elif line['col_count'] == target_col_count:
            filled_rows.append(line['columns'])

    return filled_rows


def find_winning_column_count(lines: list, min_columns: int = 2) -> int:
    """Find the most common column count (ignoring lines with < min_columns)"""
    counts = [line['col_count'] for line in lines if line['col_count'] >= min_columns]
    if not counts:
        return 0
    counter = Counter(counts)
    winner, _ = counter.most_common(1)[0]
    return winner


def extract_data_rows(lines: list) -> tuple:
    """Extract only lines matching the winning column count"""
    winner = find_winning_column_count(lines)
    data_rows = [line['columns'] for line in lines if line['col_count'] == winner]
    return data_rows, winner


def check_bank_statement(lines: list) -> dict:
    """
    Check if the PDF appears to be a bank statement.
    Returns detection info including found keywords.
    """
    if not BANKING_DETECTION_AVAILABLE:
        return {'is_bank': False, 'debit_found': [], 'credit_found': []}

    # Combine all text from lines
    all_text = ' '.join(' '.join(line['columns']) for line in lines).lower()

    # Check for debit/credit keywords
    debit_found = [kw for kw in DEBIT_KEYWORDS if kw in all_text]
    credit_found = [kw for kw in CREDIT_KEYWORDS if kw in all_text]

    is_bank = is_bank_statement(all_text)

    return {
        'is_bank': is_bank,
        'debit_found': debit_found[:3],  # Limit to 3 examples
        'credit_found': credit_found[:3],
    }


def analyze_lines(lines: list):
    """Print detailed analysis of column counts"""
    c = Colors
    counter = Counter(line['col_count'] for line in lines)

    print(f"\n{c.BOLD}Column Count Analysis{c.RESET}")
    print(f"{c.GRAY}{'─' * 50}{c.RESET}")

    winner = counter.most_common(1)[0][0] if counter else 0

    for count, freq in counter.most_common():
        bar_len = int((freq / max(counter.values())) * 20)
        bar = '█' * bar_len + '░' * (20 - bar_len)

        if count == winner:
            marker = f" {c.GREEN}← WINNER{c.RESET}"
        else:
            marker = ""

        print(f"  {c.CYAN}{count:2d} cols{c.RESET} │ {c.YELLOW}{bar}{c.RESET} │ {freq:4d} lines{marker}")

    print(f"{c.GRAY}{'─' * 50}{c.RESET}")

    # Show samples
    print(f"\n{c.BOLD}Samples by Column Count{c.RESET}\n")

    for count in sorted(counter.keys()):
        samples = [l for l in lines if l['col_count'] == count][:2]
        print(f"  {c.CYAN}{c.BOLD}{count} columns:{c.RESET}")
        for s in samples:
            preview = str(s['columns'])
            if len(preview) > 70:
                preview = preview[:67] + "..."
            print(f"    {c.DIM}{preview}{c.RESET}")
        print()


# ═══════════════════════════════════════════════════════════════════════════════
# INTERACTIVE CLI
# ═══════════════════════════════════════════════════════════════════════════════

def get_key():
    """
    Read a single keypress from stdin, handling arrow keys.

    Returns:
        'UP', 'DOWN', 'ENTER', 'QUIT', 'ESC', or the character pressed
    """
    fd = sys.stdin.fileno()
    old_settings = termios.tcgetattr(fd)
    try:
        tty.setraw(fd)
        ch = sys.stdin.read(1)

        # Handle escape sequences (arrow keys)
        if ch == '\x1b':
            ch2 = sys.stdin.read(1)
            if ch2 == '[':
                ch3 = sys.stdin.read(1)
                if ch3 == 'A':
                    return 'UP'
                elif ch3 == 'B':
                    return 'DOWN'
            return 'ESC'
        elif ch == '\r' or ch == '\n':
            return 'ENTER'
        elif ch == 'q' or ch == '\x03':  # q or Ctrl+C
            return 'QUIT'
        return ch
    finally:
        termios.tcsetattr(fd, termios.TCSADRAIN, old_settings)


def interactive_column_selector(lines: list, header_row: dict = None) -> int:
    """
    Interactive CLI for selecting column structure with arrow key navigation.

    Displays available column counts with row frequencies, allows user to
    navigate with arrow keys, and shows a live preview of selected structure.
    For bank statements, shows header row and gap-filled data preview.

    Args:
        lines: List of extracted line dicts
        header_row: Optional header dict for bank statement preview

    Returns:
        Selected column count, or 0 if cancelled
    """
    c = Colors

    # For bank statements, filter out low-column metadata rows (2-3 cols are usually key:value pairs)
    min_cols = 4 if header_row else 2
    counter = Counter(line['col_count'] for line in lines if line['col_count'] >= min_cols)

    if not counter:
        # Fallback to 2+ columns if no results
        counter = Counter(line['col_count'] for line in lines if line['col_count'] >= 2)

    if not counter:
        return 0

    # Build options grouped by OUTPUT column count (after gap-filling)
    # Maps output_col_count -> (total_rows, [source_col_counts])
    output_groups = {}
    for col_count, freq in counter.items():
        # Determine output column count
        output_cols = col_count
        # Map header-1 columns to header (gap-fill adds 1 column for Debit/Credit)
        if header_row and col_count == len(header_row['columns']) - 1:
            output_cols = len(header_row['columns'])
        # Map header-2 columns to header (gap-fill adds 2 columns: Chq + Debit/Credit)
        elif header_row and col_count == len(header_row['columns']) - 2:
            output_cols = len(header_row['columns'])  # Both 5-col and 6-col become 7-col

        if output_cols not in output_groups:
            output_groups[output_cols] = {'rows': 0, 'sources': []}
        output_groups[output_cols]['rows'] += freq
        output_groups[output_cols]['sources'].append(col_count)

    # Sort by row count (most common first)
    options = sorted(output_groups.items(), key=lambda x: -x[1]['rows'])
    total = len(options)
    selected = 0

    # Auto-select if only one option exists
    if total == 1:
        selected_count = options[0][0]
        freq = options[0][1]['rows']
        if not SILENT_MODE:
            print(f"\n  {c.GREEN}{c.BOLD}✓{c.RESET} Auto-selected: {c.WHITE}{c.BOLD}{selected_count} columns{c.RESET} ({freq} rows)\n")
        return selected_count

    # Precompute samples for each output option (with gap-filling applied)
    samples_cache = {}
    header_text = ' '.join(header_row['columns']).lower() if header_row else ''

    # Pre-check if gap-filling will work
    can_gap_fill = False
    if header_row:
        debit_idx, credit_idx = find_debit_credit_column_indices(header_row)
        can_gap_fill = debit_idx is not None and credit_idx is not None

    for output_cols, info in options:
        # Get samples from source col_counts
        # Prioritize gap-filled samples (they're the actual bank transactions)
        filled_samples = []
        other_samples = []

        for src_col in info['sources']:
            # Filter out header row from samples
            raw_samples = []
            for l in lines:
                if l['col_count'] == src_col:
                    # Skip if this looks like the header row
                    if header_row and ' '.join(l['columns']).lower() == header_text:
                        continue
                    raw_samples.append(l)
                    if len(raw_samples) >= 7:
                        break

            if header_row and can_gap_fill:
                # Try gap-filling for header-1 or header-2 column rows
                if src_col == len(header_row['columns']) - 1 or src_col == len(header_row['columns']) - 2:
                    filled = fill_empty_columns(raw_samples[:5], header_row, src_col)
                    # Verify gap-filling actually produced column count close to header
                    for f in filled:
                        if len(f) >= len(header_row['columns']) - 1:
                            filled_samples.append(f)
                        else:
                            other_samples.append(f)
                else:
                    other_samples.extend([s['columns'] for s in raw_samples[:5]])
            else:
                other_samples.extend([s['columns'] for s in raw_samples[:5]])

        # Prefer gap-filled samples, then other samples
        all_samples = filled_samples + other_samples
        samples_cache[output_cols] = all_samples[:5]

    # Calculate fixed height
    # Header(3) + separator(1) + options(1 each) + separator(1) + header(1) + samples(4) + separator(1) + controls(1) + padding(2)
    fixed_height = 3 + 1 + total + 1 + 1 + 4 + 1 + 1 + 2

    def clear_area():
        """Move cursor up and clear the display area"""
        sys.stdout.write(f"\033[{fixed_height}A")  # Move up
        sys.stdout.write("\033[J")  # Clear from cursor to end
        sys.stdout.flush()

    def render(sel_idx):
        """Render the selector UI"""
        lines_out = []

        # Header with map count
        lines_out.append("")
        lines_out.append(f"  {c.BOLD}({total}) Map found{c.RESET} - {c.DIM}Select the perfect one{c.RESET}")
        lines_out.append("")  # New line after title
        lines_out.append(f"  {c.GRAY}{'─' * 58}{c.RESET}")

        # Options list - format: (rows) columns
        for i, (output_cols, info) in enumerate(options):
            freq = info['rows']

            if i == sel_idx:
                lines_out.append(f"  {c.CYAN}{c.BOLD}▸{c.RESET} {c.GREEN}({freq} rows){c.RESET} {c.WHITE}{c.BOLD}{output_cols} columns{c.RESET}")
            else:
                lines_out.append(f"  {c.DIM}  ({freq} rows) {output_cols} columns{c.RESET}")

        lines_out.append(f"  {c.GRAY}{'─' * 58}{c.RESET}")

        # Preview of selected option
        sel_output_cols = options[sel_idx][0]
        samples = samples_cache[sel_output_cols]

        # Determine header: use bank header if available and samples match its column count
        # Check if samples actually have the expected column count (gap-filled)
        samples_have_header_cols = (
            header_row and
            samples and
            len(samples[0]) == len(header_row['columns'])
        )

        if samples_have_header_cols:
            header_cols = header_row['columns']
            preview_samples = samples  # Show all samples as data (gap-filled)
        elif samples:
            header_cols = samples[0]  # First row becomes header
            preview_samples = samples[1:] if len(samples) > 1 else []  # Rest are data
        else:
            header_cols = []
            preview_samples = []

        # Show header row (aligned with data rows)
        if header_cols:
            header_preview = ' │ '.join(str(h)[:12] for h in header_cols)
            if len(header_preview) > 50:
                header_preview = header_preview[:47] + "..."
            lines_out.append(f"  {c.GREEN}columns:{c.RESET} {c.CYAN}{header_preview}{c.RESET}")

        # Show preview rows with green row numbers (aligned with "columns:")
        for row_num, cols in enumerate(preview_samples, 1):
            preview = ' │ '.join(str(cell)[:12] for cell in cols)
            if len(preview) > 50:
                preview = preview[:47] + "..."
            lines_out.append(f"  {c.GREEN}{row_num:>7}.{c.RESET} {c.WHITE}{preview}{c.RESET}")

        lines_out.append(f"  {c.GRAY}{'─' * 58}{c.RESET}")

        # Controls at bottom with bold
        lines_out.append(f"  {c.BOLD}[↑↓] move  [Enter] select  [q] quit{c.RESET}")

        # Pad to fixed height
        while len(lines_out) < fixed_height:
            lines_out.append("")

        return '\n'.join(lines_out)

    # Hide cursor and initial render
    sys.stdout.write('\033[?25l')
    sys.stdout.flush()

    # Print initial content
    print(render(selected))

    try:
        while True:
            key = get_key()

            if key == 'UP' and selected > 0:
                selected -= 1
                clear_area()
                print(render(selected))
            elif key == 'DOWN' and selected < total - 1:
                selected += 1
                clear_area()
                print(render(selected))
            elif key == 'ENTER':
                sys.stdout.write('\033[?25h')  # Show cursor
                sys.stdout.flush()
                selected_count = options[selected][0]
                clear_area()
                print(f"\n  {c.GREEN}{c.BOLD}✓{c.RESET} Selected: {c.WHITE}{c.BOLD}{selected_count} columns{c.RESET}\n")
                return selected_count
            elif key == 'QUIT' or key == 'ESC':
                sys.stdout.write('\033[?25h')  # Show cursor
                sys.stdout.flush()
                clear_area()
                print(f"\n  {c.YELLOW}Cancelled{c.RESET}\n")
                return 0

    except Exception:
        sys.stdout.write('\033[?25h')  # Show cursor
        sys.stdout.flush()
        return 0


# ═══════════════════════════════════════════════════════════════════════════════
# DATA CLEANING & DEDUPLICATION
# ═══════════════════════════════════════════════════════════════════════════════

def deduplicate_rows(rows: list) -> tuple:
    """
    Remove duplicate rows, keeping first occurrence.

    Useful for removing repeated header rows that appear on each page.

    Args:
        rows: List of row lists

    Returns:
        Tuple of (unique_rows, count_of_duplicates_removed)
    """
    seen = set()
    unique = []
    duplicates = 0

    for row in rows:
        # Convert row to tuple for hashing
        row_tuple = tuple(row)
        if row_tuple not in seen:
            seen.add(row_tuple)
            unique.append(row)
        else:
            duplicates += 1

    return unique, duplicates


def deduplicate_columns(rows: list) -> tuple:
    """
    Remove duplicate columns where all values are identical.

    Some PDFs have redundant columns with the same data.

    Args:
        rows: List of row lists

    Returns:
        Tuple of (cleaned_rows, count_of_columns_removed)
    """
    if not rows or len(rows) < 1:
        return rows, 0

    num_cols = len(rows[0])
    if num_cols <= 1:
        return rows, 0

    # Build column data for comparison
    columns_data = []
    for col_idx in range(num_cols):
        col_values = tuple(row[col_idx] if col_idx < len(row) else '' for row in rows)
        columns_data.append(col_values)

    # Find unique columns (keep first occurrence)
    seen_columns = {}
    keep_indices = []

    for col_idx, col_data in enumerate(columns_data):
        if col_data not in seen_columns:
            seen_columns[col_data] = col_idx
            keep_indices.append(col_idx)

    columns_removed = num_cols - len(keep_indices)

    if columns_removed == 0:
        return rows, 0

    # Rebuild rows with only unique columns
    cleaned_rows = []
    for row in rows:
        new_row = [row[i] if i < len(row) else '' for i in keep_indices]
        cleaned_rows.append(new_row)

    return cleaned_rows, columns_removed


def save_csv(rows: list, output_path: str):
    """Save rows to CSV format"""
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow(row)
    return output_path


def save_tsv(rows: list, output_path: str):
    """Save rows to TSV (tab-separated) format"""
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter='\t')
        for row in rows:
            writer.writerow(row)
    return output_path


def save_json(rows: list, output_path: str):
    """Save rows to JSON format (array of arrays)"""
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)
    return output_path


def save_jsonl(rows: list, output_path: str):
    """Save rows to JSON Lines format (one JSON object per line)"""
    with open(output_path, 'w', encoding='utf-8') as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + '\n')
    return output_path


def save_markdown(rows: list, output_path: str):
    """Save rows to Markdown table format"""
    if not rows:
        return output_path

    with open(output_path, 'w', encoding='utf-8') as f:
        # Header row
        header = rows[0]
        f.write('| ' + ' | '.join(str(cell) for cell in header) + ' |\n')
        f.write('|' + '|'.join(['---' for _ in header]) + '|\n')

        # Data rows
        for row in rows[1:]:
            f.write('| ' + ' | '.join(str(cell) for cell in row) + ' |\n')

    return output_path


def save_excel(rows: list, output_path: str):
    """Save rows to Excel format (requires openpyxl)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print_error("Excel export requires openpyxl")
        print_info(f"Install with: {Colors.CYAN}pip install openpyxl{Colors.RESET}")
        print_info(f"Or use: {Colors.CYAN}pip install pdfcsv[excel]{Colors.RESET}")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    # Add rows
    for row in rows:
        ws.append(row)

    # Style header row (first row)
    if rows:
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

    wb.save(output_path)
    return output_path


# Supported output formats
OUTPUT_FORMATS = {
    'csv': {'ext': '.csv', 'func': save_csv, 'desc': 'Comma-separated values'},
    'tsv': {'ext': '.tsv', 'func': save_tsv, 'desc': 'Tab-separated values'},
    'json': {'ext': '.json', 'func': save_json, 'desc': 'JSON array'},
    'jsonl': {'ext': '.jsonl', 'func': save_jsonl, 'desc': 'JSON Lines'},
    'markdown': {'ext': '.md', 'func': save_markdown, 'desc': 'Markdown table'},
    'md': {'ext': '.md', 'func': save_markdown, 'desc': 'Markdown table'},
    'excel': {'ext': '.xlsx', 'func': save_excel, 'desc': 'Excel workbook'},
    'xlsx': {'ext': '.xlsx', 'func': save_excel, 'desc': 'Excel workbook'},
}


def print_help():
    """Print help message"""
    c = Colors
    print(f"""
{c.BOLD}USAGE{c.RESET}
    pdfcsv <pdf_file> [options]

{c.BOLD}OPTIONS{c.RESET}
    {c.CYAN}--analyze{c.RESET}        Show column distribution analysis
    {c.CYAN}--columns N{c.RESET}      Extract only lines with N columns
    {c.CYAN}--gap N{c.RESET}          Pixel gap threshold (default: 5)
    {c.CYAN}-o, --output{c.RESET}     Output filename
    {c.CYAN}-f, --format{c.RESET}     Output format (default: csv)
    {c.CYAN}-h, --help{c.RESET}       Show this help message
    {c.CYAN}-v, --version{c.RESET}    Show version
    {c.CYAN}--silent{c.RESET}         Minimal output (only result or errors)

{c.BOLD}OUTPUT FORMATS{c.RESET}
    {c.WHITE}csv{c.RESET}       Comma-separated values {c.DIM}(default){c.RESET}
    {c.WHITE}tsv{c.RESET}       Tab-separated values
    {c.WHITE}json{c.RESET}      JSON array of arrays
    {c.WHITE}jsonl{c.RESET}     JSON Lines (one object per line)
    {c.WHITE}markdown{c.RESET}  Markdown table {c.DIM}(alias: md){c.RESET}
    {c.WHITE}excel{c.RESET}     Excel workbook {c.DIM}(alias: xlsx, requires openpyxl){c.RESET}

{c.BOLD}EXAMPLES{c.RESET}
    {c.DIM}# Analyze PDF structure{c.RESET}
    pdfcsv statement.pdf --analyze

    {c.DIM}# Extract to CSV (default){c.RESET}
    pdfcsv statement.pdf --columns 6

    {c.DIM}# Export as JSON{c.RESET}
    pdfcsv statement.pdf --format json

    {c.DIM}# Export as Excel{c.RESET}
    pdfcsv statement.pdf -f excel -o transactions.xlsx

    {c.DIM}# Export as Markdown table{c.RESET}
    pdfcsv statement.pdf --format markdown

{c.BOLD}PASSWORD-PROTECTED PDFs{c.RESET}
    If a PDF is password-protected, you'll be prompted to enter
    the password. Enter 'q' or press Ctrl+C to cancel.

{c.BOLD}HOW IT WORKS{c.RESET}
    1. Extracts characters with pixel positions from PDF
    2. Groups characters into lines by Y coordinate
    3. Detects columns by measuring pixel gaps
    4. Counts columns per line
    5. Exports rows matching target column count

{c.DIM}GitHub: https://github.com/stexz01/pdfcsv{c.RESET}
""")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    c = Colors

    # Show version
    if '--version' in sys.argv or '-v' in sys.argv:
        print(f"PDFcsv v{__version__} by {__author__}")
        sys.exit(0)

    # Show help
    if len(sys.argv) < 2 or '--help' in sys.argv or '-h' in sys.argv:
        banner()
        print_help()
        sys.exit(0)

    # Parse arguments
    pdf_path = sys.argv[1]
    analyze_mode = '--analyze' in sys.argv

    # Check for --silent mode
    global SILENT_MODE
    SILENT_MODE = '--silent' in sys.argv

    # Check for --columns override (support both --columns and --column)
    target_columns = None
    if '--columns' in sys.argv or '--column' in sys.argv:
        flag = '--columns' if '--columns' in sys.argv else '--column'
        idx = sys.argv.index(flag)
        if idx + 1 < len(sys.argv):
            target_columns = int(sys.argv[idx + 1])

    # Check for --gap override
    gap_threshold = 5
    if '--gap' in sys.argv:
        idx = sys.argv.index('--gap')
        if idx + 1 < len(sys.argv):
            gap_threshold = int(sys.argv[idx + 1])

    # Check for --output override
    output_path = None
    if '--output' in sys.argv or '-o' in sys.argv:
        flag = '--output' if '--output' in sys.argv else '-o'
        idx = sys.argv.index(flag)
        if idx + 1 < len(sys.argv):
            output_path = sys.argv[idx + 1]

    # Check for --format override
    output_format = 'csv'
    if '--format' in sys.argv or '-f' in sys.argv:
        flag = '--format' if '--format' in sys.argv else '-f'
        idx = sys.argv.index(flag)
        if idx + 1 < len(sys.argv):
            output_format = sys.argv[idx + 1].lower()
            if output_format not in OUTPUT_FORMATS:
                print_error(f"Unknown format: {output_format}")
                print_info(f"Supported: {', '.join(sorted(set(f for f in OUTPUT_FORMATS.keys() if len(f) > 2)))}")
                sys.exit(1)

    # Banner (skip in silent mode)
    if not SILENT_MODE:
        banner()

    # Validate input file
    if not os.path.exists(pdf_path):
        print_error(f"File not found: {pdf_path}")
        sys.exit(1)

    # Step 1: Extract (try table extraction first, fall back to gap-based)
    print_step(1, f"Reading {c.WHITE}{c.BOLD}{pdf_path}{c.RESET}")

    is_table_mode = False
    with open_pdf(pdf_path) as pdf:
        # Try structured table extraction first (handles multi-line cells)
        raw_table_rows = _extract_tables(pdf)
        table_rows = clean_table_rows(raw_table_rows)

        if table_rows and len(table_rows) >= 3 and not _table_has_concatenated_cells(table_rows):
            lines = table_rows_to_lines(table_rows)
            is_table_mode = True
            print_success(f"Structured tables detected ({c.WHITE}{len(lines)}{c.RESET} rows)")

            # Warn if column headers were unextractable (vector-drawn text)
            _warn_broken_headers(raw_table_rows, table_rows)
        else:
            # Fallback: gap-based character extraction
            print_info(f"Gap threshold: {gap_threshold}px")
            lines = _extract_gaps(pdf, gap_threshold)
            print_success(f"Extracted {c.WHITE}{len(lines)}{c.RESET} lines")

    # Check if this is a bank statement and find header early (for preview)
    bank_info = check_bank_statement(lines)
    header_row = None
    if bank_info['is_bank']:
        if is_table_mode:
            print_info(f"Bank statement detected (table extraction handles column alignment)")
        else:
            print_warning(f"{c.YELLOW}Bank statement detected{c.RESET}")
            print_info(f"{c.DIM}Empty debit/credit columns may not appear in output{c.RESET}")
            if bank_info['debit_found']:
                print_info(f"{c.DIM}Debit keywords found: {', '.join(bank_info['debit_found'][:3])}{c.RESET}")
            if bank_info['credit_found']:
                print_info(f"{c.DIM}Credit keywords found: {', '.join(bank_info['credit_found'][:3])}{c.RESET}")

            # Find header row for gap-filling preview
            header_row = find_header_row(lines)

    # Step 2: Analyze (optional)
    if analyze_mode:
        print_step(2, "Analyzing column structure")
        analyze_lines(lines)

        if not target_columns:
            print_info(f"Run with {c.CYAN}--columns N{c.RESET} to extract specific column count")
            sys.exit(0)

    # Step 3: Extract data rows
    step_num = 3 if analyze_mode else 2

    if target_columns:
        winner = target_columns
        print_step(step_num, f"Filtering for {c.WHITE}{winner}{c.RESET} column rows")
    elif not analyze_mode:
        # Interactive mode - let user pick (pass header for preview with gap-filling)
        print()  # Add spacing before interactive selector
        print()
        winner = interactive_column_selector(lines, header_row)
        if winner == 0:
            print_info("No selection made. Exiting.")
            sys.exit(0)
        step_num += 1
    else:
        winner = find_winning_column_count(lines)
        print_step(step_num, f"Auto-detected {c.WHITE}{winner}{c.RESET} columns as target")

    # Apply gap-filling for bank statements (only for gap-based extraction)
    if not is_table_mode and bank_info['is_bank'] and header_row:
        debit_idx, credit_idx = find_debit_credit_column_indices(header_row)
        header_col_count = len(header_row['columns'])

        if debit_idx is not None and credit_idx is not None:
            # Gap-fill and collect all rows that result in winner column count
            filled = fill_empty_columns(lines, header_row, winner)

            # Filter to only include rows matching winner count
            data_rows = [row for row in filled if len(row) == winner]

            # If no exact matches, also check for header-1 match (gap-filled from header-2)
            if not data_rows and winner == header_col_count:
                # Include rows that are close to header count
                data_rows = [row for row in filled if len(row) >= header_col_count - 1]

            if data_rows:
                print_success(f"Gap-filling applied for empty debit/credit columns")
            else:
                # Fallback to raw data if gap-filling didn't produce matching rows
                data_rows = [line['columns'] for line in lines if line['col_count'] == winner]
        else:
            data_rows = [line['columns'] for line in lines if line['col_count'] == winner]
    else:
        data_rows = [line['columns'] for line in lines if line['col_count'] == winner]

    if not data_rows:
        print_warning("No matching rows found")
        print_info(f"Try {c.CYAN}--analyze{c.RESET} to see column distribution")
        sys.exit(1)

    # Deduplicate rows (remove repeated headers, etc.)
    data_rows, rows_removed = deduplicate_rows(data_rows)

    # Deduplicate columns (remove identical columns)
    data_rows, cols_removed = deduplicate_columns(data_rows)

    # Build status message
    dedup_info = []
    if rows_removed > 0:
        dedup_info.append(f"{rows_removed} duplicate rows")
    if cols_removed > 0:
        dedup_info.append(f"{cols_removed} duplicate cols")

    if dedup_info:
        print_success(f"Found {c.WHITE}{len(data_rows)}{c.RESET} data rows {c.DIM}(removed: {', '.join(dedup_info)}){c.RESET}")
    else:
        print_success(f"Found {c.WHITE}{len(data_rows)}{c.RESET} data rows")

    # Step 4: Save
    step_num += 1
    format_info = OUTPUT_FORMATS[output_format]
    if not output_path:
        output_path = pdf_path.rsplit('.', 1)[0] + '_extracted' + format_info['ext']

    print_step(step_num, f"Saving to {c.WHITE}{output_path}{c.RESET} {c.DIM}({format_info['desc']}){c.RESET}")
    save_func = format_info['func']
    actual_path = save_func(data_rows, output_path)
    print_success(f"Saved {c.GREEN}{c.BOLD}{actual_path}{c.RESET}")

    # Preview (skip in silent mode)
    if not SILENT_MODE:
        print(f"\n{c.BOLD}Preview (first 5 rows){c.RESET}")
        print(f"{c.GRAY}{'─' * 50}{c.RESET}")

        # Show header row for bank statements if columns match
        if bank_info['is_bank'] and header_row and data_rows:
            if len(data_rows[0]) == len(header_row['columns']):
                header_preview = ' │ '.join(str(h)[:15] for h in header_row['columns'])
                print(f"  {c.GREEN}columns:{c.RESET} {c.CYAN}{header_preview}{c.RESET}")

        for idx, row in enumerate(data_rows[:5], 1):
            preview = ' │ '.join(str(cell)[:15] for cell in row)
            print(f"  {c.GREEN}{idx:>7}.{c.RESET} {c.WHITE}{preview}{c.RESET}")
        print(f"{c.GRAY}{'─' * 50}{c.RESET}")

    # Done (always show, simpler in silent mode)
    if SILENT_MODE:
        print(f"Done! Extracted {len(data_rows)} rows to {actual_path}")
    else:
        print(f"\n{c.GREEN}{c.BOLD}Done!{c.RESET} {c.DIM}Extracted {len(data_rows)} rows to {actual_path}{c.RESET}\n")


if __name__ == "__main__":
    main()
